from fastapi import FastAPI, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, DateTime, ForeignKey, and_, extract
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from datetime import datetime, date
import os
import io
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Tạo database
SQLALCHEMY_DATABASE_URL = "sqlite:///./transport.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Tạo templates với custom filters
templates = Jinja2Templates(directory="templates")

# Thêm custom filter để parse JSON
def from_json(value):
    import json
    try:
        return json.loads(value) if value else []
    except:
        return []

# Đăng ký filter
templates.env.filters["from_json"] = from_json

# Models
class Employee(Base):
    __tablename__ = "employees"
    
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    birth_date = Column(Date)  # Ngày tháng năm sinh
    phone = Column(String)
    cccd = Column(String)  # Số CCCD
    cccd_issue_date = Column(Date)  # Ngày cấp CCCD
    cccd_expiry = Column(Date)  # Ngày hết hạn CCCD
    driving_license = Column(String)  # Số bằng lái xe
    license_expiry = Column(Date)  # Ngày hết hạn bằng lái
    documents = Column(String)  # Đường dẫn file upload giấy tờ (JSON array)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships removed - no longer linked to routes

class Vehicle(Base):
    __tablename__ = "vehicles"
    
    id = Column(Integer, primary_key=True, index=True)
    license_plate = Column(String, unique=True, nullable=False)
    vehicle_info = Column(String)  # Thông tin xe (model/loại)
    capacity = Column(Float)  # Trọng tải
    fuel_consumption = Column(Float)  # Tiêu hao nhiên liệu
    inspection_expiry = Column(Date)  # Ngày hết hạn đăng kiểm
    inspection_documents = Column(String)  # Đường dẫn file upload sổ đăng kiểm (JSON array)
    phu_hieu_expired_date = Column(Date)  # Ngày hết hạn phù hiệu vận tải
    phu_hieu_files = Column(String)  # Đường dẫn file upload phù hiệu vận tải (JSON array)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    routes = relationship("Route", back_populates="vehicle")

class Route(Base):
    __tablename__ = "routes"
    
    id = Column(Integer, primary_key=True, index=True)
    route_code = Column(String, nullable=False)  # NA_002, NA_004, etc.
    route_name = Column(String, nullable=False)
    distance = Column(Float)  # KM/Chuyến
    unit_price = Column(Float)  # Đơn giá
    monthly_salary = Column(Float)  # Lương tuyến/tháng
    vehicle_id = Column(Integer, ForeignKey("vehicles.id"), nullable=True)
    is_active = Column(Integer, default=1)
    status = Column(Integer, default=1)  # 1: Active, 0: Inactive
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", back_populates="routes")
    daily_routes = relationship("DailyRoute", back_populates="route")

class DailyRoute(Base):
    __tablename__ = "daily_routes"
    
    id = Column(Integer, primary_key=True, index=True)
    route_id = Column(Integer, ForeignKey("routes.id"))
    date = Column(Date, nullable=False)
    distance_km = Column(Float, default=0)  # Số km
    cargo_weight = Column(Float, default=0)  # Tải trọng
    driver_name = Column(String)  # Tên lái xe
    license_plate = Column(String)  # Biển số xe
    employee_name = Column(String)  # Tên nhân viên
    notes = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    route = relationship("Route", back_populates="daily_routes")

class FuelRecord(Base):
    __tablename__ = "fuel_records"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date, nullable=False)  # Ngày đổ dầu
    fuel_type = Column(String, default="Dầu DO 0,05S-II")  # Loại dầu
    license_plate = Column(String, nullable=False)  # Biển số xe
    fuel_price_per_liter = Column(Float, default=0)  # Giá xăng dầu hôm nay (đồng/lít)
    liters_pumped = Column(Float, default=0)  # Số lít dầu đã đổ
    cost_pumped = Column(Float, default=0)  # Số tiền dầu đã đổ (tự động tính)
    notes = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)
    
    # Relationships
    vehicle = relationship("Vehicle", foreign_keys=[license_plate], primaryjoin="FuelRecord.license_plate == Vehicle.license_plate")

class FinanceRecord(Base):
    __tablename__ = "finance_records"
    
    id = Column(Integer, primary_key=True, index=True)
    date = Column(Date, nullable=False)  # Ngày giao dịch
    category = Column(String, nullable=False)  # Danh mục (Thu/Chi)
    description = Column(String, nullable=False)  # Diễn giải/Tên khách hàng
    route_code = Column(String)  # Mã tuyến
    amount_before_vat = Column(Float, default=0)  # Số tiền (chưa VAT)
    vat_rate = Column(Float, default=0)  # VAT (%)
    discount1_rate = Column(Float, default=0)  # Chiết khấu 1 (%)
    discount2_rate = Column(Float, default=0)  # Chiết khấu 2 (%)
    final_amount = Column(Float, default=0)  # Thành tiền (tự động tính)
    income = Column(Float, default=0)  # Số tiền thu (để tương thích)
    expense = Column(Float, default=0)  # Số tiền chi (để tương thích)
    balance = Column(Float, default=0)  # Thành tiền (tự động tính)
    notes = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)

class FinanceTransaction(Base):
    """Bảng riêng biệt chuyên quản lý dữ liệu thu chi độc lập"""
    __tablename__ = "finance_transactions"
    
    id = Column(Integer, primary_key=True, index=True)
    transaction_type = Column(String, nullable=False)  # Thu/Chi
    category = Column(String, nullable=False)  # Danh mục
    date = Column(Date, nullable=False)  # Ngày thu/chi
    description = Column(String, nullable=False)  # Diễn giải
    route_code = Column(String)  # Mã tuyến (nếu có)
    amount = Column(Float, default=0)  # Số tiền chưa VAT
    vat = Column(Float, default=0)  # VAT (%)
    discount1 = Column(Float, default=0)  # Chiết khấu 1 (%)
    discount2 = Column(Float, default=0)  # Chiết khấu 2 (%)
    total = Column(Float, default=0)  # Thành tiền
    note = Column(String)  # Ghi chú
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Tạo bảng
Base.metadata.create_all(bind=engine)

# Dependency để lấy database session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# FastAPI app
app = FastAPI(title="Hệ thống quản lý vận chuyển")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates đã được tạo ở trên với custom filters

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    # Lấy thống kê tổng quan
    employees_count = db.query(Employee).count()
    vehicles_count = db.query(Vehicle).count()
    routes_count = db.query(Route).filter(Route.is_active == 1).count()
    today = date.today()
    daily_routes_count = db.query(DailyRoute).filter(DailyRoute.date == today).count()
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "employees_count": employees_count,
        "vehicles_count": vehicles_count,
        "routes_count": routes_count,
        "daily_routes_count": daily_routes_count
    })

@app.get("/report", response_class=HTMLResponse)
async def report_page(request: Request):
    """Trang báo cáo tổng hợp - menu chính cho các báo cáo"""
    return templates.TemplateResponse("report.html", {"request": request})

@app.get("/employees", response_class=HTMLResponse)
async def employees_page(request: Request, db: Session = Depends(get_db)):
    employees = db.query(Employee).filter(Employee.status == 1).all()
    return templates.TemplateResponse("employees.html", {"request": request, "employees": employees})


@app.get("/employees/documents/{employee_id}")
async def get_employee_documents(employee_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin giấy tờ của nhân viên"""
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy nhân viên"}
        )
    
    if not employee.documents:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Nhân viên chưa upload giấy tờ"}
        )
    
    try:
        import json
        documents = json.loads(employee.documents)
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            file_path = f"static/uploads/{doc}"
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                file_extension = os.path.splitext(doc)[1].lower()
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "size": file_size,
                    "extension": file_extension,
                    "exists": True
                })
            else:
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "exists": False
                })
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError:
        # Xử lý dữ liệu cũ (không phải JSON)
        if isinstance(employee.documents, str) and employee.documents.strip():
            file_path = f"static/uploads/{employee.documents}"
            if os.path.exists(file_path):
                return JSONResponse(
                    status_code=200,
                    content={
                        "success": True,
                        "documents": [{
                            "filename": employee.documents,
                            "url": f"/static/uploads/{employee.documents}",
                            "size": os.path.getsize(file_path),
                            "extension": os.path.splitext(employee.documents)[1].lower(),
                            "exists": True
                        }],
                        "total": 1
                    }
                )
        
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Dữ liệu giấy tờ không hợp lệ"}
        )

@app.post("/employees/add")
async def add_employee(
    name: str = Form(...),
    birth_date: str = Form(""),
    phone: str = Form(""),
    cccd: str = Form(""),
    cccd_issue_date: str = Form(""),
    cccd_expiry: str = Form(""),
    driving_license: str = Form(""),
    license_expiry: str = Form(""),
    documents: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    # Convert date strings to date objects
    birth_date_obj = None
    cccd_issue_date_obj = None
    cccd_expiry_date = None
    license_expiry_date = None
    
    if birth_date:
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()
    if cccd_issue_date:
        cccd_issue_date_obj = datetime.strptime(cccd_issue_date, "%Y-%m-%d").date()
    if cccd_expiry:
        cccd_expiry_date = datetime.strptime(cccd_expiry, "%Y-%m-%d").date()
    if license_expiry:
        license_expiry_date = datetime.strptime(license_expiry, "%Y-%m-%d").date()
    
    # Handle multiple file uploads
    documents_paths = []
    if documents:
        for document in documents:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                documents_paths.append(filename)
    
    # Convert documents list to JSON string
    documents_json = json.dumps(documents_paths) if documents_paths else None
    
    employee = Employee(
        name=name,
        birth_date=birth_date_obj,
        phone=phone, 
        cccd=cccd,
        cccd_issue_date=cccd_issue_date_obj,
        cccd_expiry=cccd_expiry_date,
        driving_license=driving_license,
        license_expiry=license_expiry_date,
        documents=documents_json
    )
    db.add(employee)
    db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.post("/employees/delete/{employee_id}")
async def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if employee:
        employee.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.get("/employees/edit/{employee_id}", response_class=HTMLResponse)
async def edit_employee_page(request: Request, employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return RedirectResponse(url="/employees", status_code=303)
    return templates.TemplateResponse("edit_employee.html", {"request": request, "employee": employee})

@app.post("/employees/edit/{employee_id}")
async def edit_employee(
    employee_id: int,
    name: str = Form(...),
    birth_date: str = Form(""),
    phone: str = Form(""),
    cccd: str = Form(""),
    cccd_issue_date: str = Form(""),
    cccd_expiry: str = Form(""),
    driving_license: str = Form(""),
    license_expiry: str = Form(""),
    documents: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return RedirectResponse(url="/employees", status_code=303)
    
    # Convert date strings to date objects
    birth_date_obj = None
    cccd_issue_date_obj = None
    cccd_expiry_date = None
    license_expiry_date = None
    
    if birth_date:
        birth_date_obj = datetime.strptime(birth_date, "%Y-%m-%d").date()
    if cccd_issue_date:
        cccd_issue_date_obj = datetime.strptime(cccd_issue_date, "%Y-%m-%d").date()
    if cccd_expiry:
        cccd_expiry_date = datetime.strptime(cccd_expiry, "%Y-%m-%d").date()
    if license_expiry:
        license_expiry_date = datetime.strptime(license_expiry, "%Y-%m-%d").date()
    
    # Handle multiple file uploads
    if documents:
        documents_paths = []
        for document in documents:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                documents_paths.append(filename)
        
        if documents_paths:
            employee.documents = json.dumps(documents_paths)
    
    # Update employee data
    employee.name = name
    employee.birth_date = birth_date_obj
    employee.phone = phone
    employee.cccd = cccd
    employee.cccd_issue_date = cccd_issue_date_obj
    employee.cccd_expiry = cccd_expiry_date
    employee.driving_license = driving_license
    employee.license_expiry = license_expiry_date
    
    db.commit()
    return RedirectResponse(url="/employees", status_code=303)

@app.delete("/employees/documents/{employee_id}")
async def delete_employee_document(
    employee_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa giấy tờ của nhân viên"""
    employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
    if not employee:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy nhân viên"}
        )
    
    if not employee.documents:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Nhân viên chưa có giấy tờ nào"}
        )
    
    try:
        import json
        documents = json.loads(employee.documents)
        
        # Kiểm tra file có tồn tại trong danh sách không
        if filename not in documents:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách giấy tờ"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        file_path = f"static/uploads/{filename}"
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                # Log lỗi nhưng vẫn tiếp tục xóa khỏi DB
                print(f"Lỗi khi xóa file {file_path}: {str(e)}")
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(filename)
        
        if documents:
            # Còn giấy tờ khác, cập nhật danh sách
            employee.documents = json.dumps(documents)
        else:
            # Không còn giấy tờ nào, set null
            employee.documents = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa giấy tờ thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu giấy tờ"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.get("/vehicles", response_class=HTMLResponse)
async def vehicles_page(request: Request, db: Session = Depends(get_db)):
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    return templates.TemplateResponse("vehicles.html", {"request": request, "vehicles": vehicles, "today": today})

@app.post("/vehicles/add")
async def add_vehicle(
    license_plate: str = Form(...),
    vehicle_info: str = Form(""),
    capacity: float = Form(0),
    fuel_consumption: float = Form(0),
    inspection_expiry: str = Form(""),
    inspection_documents: list[UploadFile] = File(None),
    phu_hieu_expired_date: str = Form(""),
    phu_hieu_files: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    # Convert date string to date object
    inspection_expiry_date = None
    if inspection_expiry:
        try:
            inspection_expiry_date = datetime.strptime(inspection_expiry, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle multiple file uploads
    documents_paths = []
    if inspection_documents:
        for document in inspection_documents:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                documents_paths.append(filename)
    
    # Convert documents list to JSON string
    documents_json = json.dumps(documents_paths) if documents_paths else None
    
    # Handle phù hiệu vận tải date
    phu_hieu_expired_date_obj = None
    if phu_hieu_expired_date:
        try:
            phu_hieu_expired_date_obj = datetime.strptime(phu_hieu_expired_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle phù hiệu vận tải file uploads
    phu_hieu_paths = []
    if phu_hieu_files:
        for document in phu_hieu_files:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                phu_hieu_paths.append(filename)
    
    # Convert phù hiệu files list to JSON string
    phu_hieu_json = json.dumps(phu_hieu_paths) if phu_hieu_paths else None
    
    vehicle = Vehicle(
        license_plate=license_plate,
        vehicle_info=vehicle_info,
        capacity=capacity,
        fuel_consumption=fuel_consumption,
        inspection_expiry=inspection_expiry_date,
        inspection_documents=documents_json,
        phu_hieu_expired_date=phu_hieu_expired_date_obj,
        phu_hieu_files=phu_hieu_json
    )
    db.add(vehicle)
    db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.post("/vehicles/delete/{vehicle_id}")
async def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if vehicle:
        vehicle.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.get("/vehicles/edit/{vehicle_id}", response_class=HTMLResponse)
async def edit_vehicle_page(request: Request, vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return RedirectResponse(url="/vehicles", status_code=303)
    return templates.TemplateResponse("edit_vehicle.html", {"request": request, "vehicle": vehicle})

@app.post("/vehicles/edit/{vehicle_id}")
async def edit_vehicle(
    vehicle_id: int,
    license_plate: str = Form(...),
    vehicle_info: str = Form(""),
    capacity: float = Form(0),
    fuel_consumption: float = Form(0),
    inspection_expiry: str = Form(""),
    inspection_documents: list[UploadFile] = File(None),
    phu_hieu_expired_date: str = Form(""),
    phu_hieu_files: list[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    import json
    
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return RedirectResponse(url="/vehicles", status_code=303)
    
    # Convert date string to date object
    inspection_expiry_date = None
    if inspection_expiry:
        try:
            inspection_expiry_date = datetime.strptime(inspection_expiry, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle multiple file uploads - append to existing documents
    if inspection_documents:
        documents_paths = []
        for document in inspection_documents:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                documents_paths.append(filename)
        
        if documents_paths:
            # Get existing documents and append new ones
            existing_documents = []
            if vehicle.inspection_documents:
                try:
                    existing_documents = json.loads(vehicle.inspection_documents)
                except json.JSONDecodeError:
                    existing_documents = []
            
            # Combine existing and new documents
            all_documents = existing_documents + documents_paths
            vehicle.inspection_documents = json.dumps(all_documents)
    
    # Handle phù hiệu vận tải date
    phu_hieu_expired_date_obj = None
    if phu_hieu_expired_date:
        try:
            phu_hieu_expired_date_obj = datetime.strptime(phu_hieu_expired_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    # Handle phù hiệu vận tải file uploads - append to existing files
    if phu_hieu_files:
        phu_hieu_paths = []
        for document in phu_hieu_files:
            if document and document.filename:
                # Validate file type
                allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif']
                file_extension = os.path.splitext(document.filename)[1].lower()
                
                if file_extension not in allowed_extensions:
                    continue  # Skip invalid files
                
                # Create unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{document.filename}"
                file_path = f"static/uploads/{filename}"
                
                # Save file
                with open(file_path, "wb") as buffer:
                    content = await document.read()
                    buffer.write(content)
                
                phu_hieu_paths.append(filename)
        
        if phu_hieu_paths:
            # Get existing phù hiệu files and append new ones
            existing_phu_hieu = []
            if vehicle.phu_hieu_files:
                try:
                    existing_phu_hieu = json.loads(vehicle.phu_hieu_files)
                except json.JSONDecodeError:
                    existing_phu_hieu = []
            
            # Combine existing and new phù hiệu files
            all_phu_hieu = existing_phu_hieu + phu_hieu_paths
            vehicle.phu_hieu_files = json.dumps(all_phu_hieu)
    
    # Update vehicle data
    vehicle.license_plate = license_plate
    vehicle.vehicle_info = vehicle_info
    vehicle.capacity = capacity
    vehicle.fuel_consumption = fuel_consumption
    vehicle.inspection_expiry = inspection_expiry_date
    vehicle.phu_hieu_expired_date = phu_hieu_expired_date_obj
    
    db.commit()
    return RedirectResponse(url="/vehicles", status_code=303)

@app.get("/vehicles/documents/{vehicle_id}")
async def get_vehicle_documents(vehicle_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin sổ đăng kiểm của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.inspection_documents:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Xe chưa upload sổ đăng kiểm"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.inspection_documents)
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            file_path = f"static/uploads/{doc}"
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                file_extension = os.path.splitext(doc)[1].lower()
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "size": file_size,
                    "extension": file_extension,
                    "exists": True
                })
            else:
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "exists": False
                })
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Dữ liệu sổ đăng kiểm không hợp lệ"}
        )

@app.delete("/vehicles/documents/{vehicle_id}")
async def delete_vehicle_document(
    vehicle_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa sổ đăng kiểm của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.inspection_documents:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Xe chưa có sổ đăng kiểm nào"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.inspection_documents)
        
        # Kiểm tra file có tồn tại trong danh sách không
        if filename not in documents:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách sổ đăng kiểm"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        file_path = f"static/uploads/{filename}"
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                # Log lỗi nhưng vẫn tiếp tục xóa khỏi DB
                print(f"Lỗi khi xóa file {file_path}: {str(e)}")
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(filename)
        
        if documents:
            # Còn sổ đăng kiểm khác, cập nhật danh sách
            vehicle.inspection_documents = json.dumps(documents)
        else:
            # Không còn sổ đăng kiểm nào, set null
            vehicle.inspection_documents = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa sổ đăng kiểm thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu sổ đăng kiểm"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.get("/vehicles/phu-hieu-documents/{vehicle_id}")
async def get_vehicle_phu_hieu_documents(vehicle_id: int, db: Session = Depends(get_db)):
    """API để lấy thông tin phù hiệu vận tải của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.phu_hieu_files:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Xe chưa upload phù hiệu vận tải"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.phu_hieu_files)
        
        # Kiểm tra file tồn tại
        existing_documents = []
        for doc in documents:
            file_path = f"static/uploads/{doc}"
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                file_extension = os.path.splitext(doc)[1].lower()
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "size": file_size,
                    "extension": file_extension,
                    "exists": True
                })
            else:
                existing_documents.append({
                    "filename": doc,
                    "url": f"/static/uploads/{doc}",
                    "exists": False
                })
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "documents": existing_documents,
                "total": len(existing_documents)
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=200,
            content={"success": True, "documents": [], "message": "Dữ liệu phù hiệu vận tải không hợp lệ"}
        )

@app.delete("/vehicles/phu-hieu-documents/{vehicle_id}")
async def delete_vehicle_phu_hieu_document(
    vehicle_id: int, 
    filename: str,
    db: Session = Depends(get_db)
):
    """API để xóa phù hiệu vận tải của xe"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.status == 1).first()
    if not vehicle:
        return JSONResponse(
            status_code=404,
            content={"success": False, "error": "Không tìm thấy xe"}
        )
    
    if not vehicle.phu_hieu_files:
        return JSONResponse(
            status_code=400,
            content={"success": False, "error": "Xe chưa có phù hiệu vận tải nào"}
        )
    
    try:
        import json
        documents = json.loads(vehicle.phu_hieu_files)
        
        # Kiểm tra file có tồn tại trong danh sách không
        if filename not in documents:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "File không tồn tại trong danh sách phù hiệu vận tải"}
            )
        
        # Xóa file khỏi thư mục lưu trữ
        file_path = f"static/uploads/{filename}"
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                # Log lỗi nhưng vẫn tiếp tục xóa khỏi DB
                print(f"Lỗi khi xóa file {file_path}: {str(e)}")
        
        # Xóa file khỏi danh sách trong DB
        documents.remove(filename)
        
        if documents:
            # Còn phù hiệu vận tải khác, cập nhật danh sách
            vehicle.phu_hieu_files = json.dumps(documents)
        else:
            # Không còn phù hiệu vận tải nào, set null
            vehicle.phu_hieu_files = None
        
        db.commit()
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True, 
                "message": "Xóa phù hiệu vận tải thành công",
                "remaining_documents": len(documents) if documents else 0
            }
        )
        
    except json.JSONDecodeError:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Lỗi định dạng dữ liệu phù hiệu vận tải"}
        )
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": f"Lỗi hệ thống: {str(e)}"}
        )

@app.get("/routes", response_class=HTMLResponse)
async def routes_page(request: Request, db: Session = Depends(get_db)):
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    return templates.TemplateResponse("routes.html", {
        "request": request, 
        "routes": routes
    })

@app.post("/routes/add")
async def add_route(
    route_code: str = Form(...),
    route_name: str = Form(...),
    distance: float = Form(0),
    monthly_salary: float = Form(0),
    db: Session = Depends(get_db)
):
    route = Route(
        route_code=route_code,
        route_name=route_name,
        distance=distance,
        unit_price=0,  # Set to 0 as default since field is removed from form
        monthly_salary=monthly_salary,
        vehicle_id=None  # No vehicle assigned by default
    )
    db.add(route)
    db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.post("/routes/delete/{route_id}")
async def delete_route(route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if route:
        route.status = 0  # Soft delete
        db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.get("/routes/edit/{route_id}", response_class=HTMLResponse)
async def edit_route_page(request: Request, route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if not route:
        return RedirectResponse(url="/routes", status_code=303)
    return templates.TemplateResponse("edit_route.html", {
        "request": request, 
        "route": route
    })

@app.post("/routes/edit/{route_id}")
async def edit_route(
    route_id: int,
    route_code: str = Form(...),
    route_name: str = Form(...),
    distance: float = Form(0),
    monthly_salary: float = Form(0),
    db: Session = Depends(get_db)
):
    route = db.query(Route).filter(Route.id == route_id, Route.status == 1).first()
    if not route:
        return RedirectResponse(url="/routes", status_code=303)
    
    route.route_code = route_code
    route.route_name = route_name
    route.distance = distance
    # unit_price is not updated since field is removed from form
    route.monthly_salary = monthly_salary
    
    db.commit()
    return RedirectResponse(url="/routes", status_code=303)

@app.get("/daily", response_class=HTMLResponse)
async def daily_page(request: Request, db: Session = Depends(get_db), selected_date: Optional[str] = None):
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    
    # Xử lý ngày được chọn
    print(f"DEBUG: selected_date parameter = {selected_date}")
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            print(f"DEBUG: Parsed filter_date = {filter_date}")
        except ValueError:
            print(f"DEBUG: Invalid date format, using today")
            filter_date = today
    else:
        print(f"DEBUG: No selected_date, using today")
        filter_date = today
    
    # Lọc chuyến đã ghi nhận theo ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == filter_date).order_by(DailyRoute.created_at.desc()).all()
    
    # Debug: Print to console
    print(f"DEBUG: Routes count: {len(routes)}")
    print(f"DEBUG: Employees count: {len(employees)}")
    print(f"DEBUG: Vehicles count: {len(vehicles)}")
    print(f"DEBUG: Filter date: {filter_date}")
    print(f"DEBUG: Daily routes count: {len(daily_routes)}")
    if vehicles:
        for v in vehicles:
            print(f"DEBUG: Vehicle: {v.license_plate} (ID: {v.id}, Status: {v.status})")
    else:
        print("DEBUG: No vehicles found!")
        # Check all vehicles regardless of status
        all_vehicles = db.query(Vehicle).all()
        print(f"DEBUG: Total vehicles in DB: {len(all_vehicles)}")
        for v in all_vehicles:
            print(f"DEBUG: All Vehicle: {v.license_plate} (ID: {v.id}, Status: {v.status})")
    
    return templates.TemplateResponse("daily.html", {
        "request": request,
        "routes": routes,
        "employees": employees,
        "vehicles": vehicles,
        "daily_routes": daily_routes,
        "today": today,
        "selected_date": filter_date.strftime('%Y-%m-%d'),
        "filter_date": filter_date
    })

@app.post("/daily/add")
async def add_daily_route(request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    
    # Lấy ngày được chọn từ form
    selected_date_str = form_data.get("date")
    if not selected_date_str:
        return RedirectResponse(url="/daily", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = date.today()
    
    # Lấy tất cả routes
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Xử lý từng route
    for route in routes:
        route_id = route.id
        
        # Lấy dữ liệu từ form cho route này
        distance_km = form_data.get(f"distance_km_{route_id}")
        driver_name = form_data.get(f"driver_name_{route_id}")
        license_plate = form_data.get(f"license_plate_{route_id}")
        notes = form_data.get(f"notes_{route_id}")
        
        # Chỉ tạo record nếu có ít nhất một trường được điền
        if distance_km or driver_name or license_plate or notes:
            daily_route = DailyRoute(
                route_id=route_id,
                date=selected_date,
                distance_km=float(distance_km) if distance_km else 0,
                cargo_weight=0,  # Set default value
                driver_name=driver_name or "",
                license_plate=license_plate or "",
                employee_name="",  # Empty since we removed this field
                notes=notes or ""
            )
            db.add(daily_route)
    
    db.commit()
    # Redirect về trang daily với ngày đã chọn
    return RedirectResponse(url=f"/daily?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.post("/daily/delete/{daily_route_id}")
async def delete_daily_route(daily_route_id: int, request: Request, db: Session = Depends(get_db)):
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if daily_route:
        # Lưu ngày của chuyến bị xóa để redirect về đúng ngày
        deleted_date = daily_route.date
        db.delete(daily_route)
        db.commit()
        return RedirectResponse(url=f"/daily?selected_date={deleted_date.strftime('%Y-%m-%d')}", status_code=303)
    return RedirectResponse(url="/daily", status_code=303)

# New Daily Page with simple date selection
@app.get("/daily-new", response_class=HTMLResponse)
async def daily_new_page(request: Request, db: Session = Depends(get_db), selected_date: Optional[str] = None, deleted_all: Optional[str] = None):
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    today = date.today()
    
    # Xử lý ngày được chọn
    if selected_date:
        try:
            filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            filter_date = today
    else:
        filter_date = today
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Sắp xếp employees theo tên (A-Z) để dễ tìm kiếm trong dropdown
    employees = sorted(employees, key=lambda emp: emp.name.lower() if emp.name else "")
    
    # Lọc chuyến đã ghi nhận theo ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == filter_date).order_by(DailyRoute.created_at.desc()).all()
    
    return templates.TemplateResponse("daily_new.html", {
        "request": request,
        "routes": routes,
        "employees": employees,
        "vehicles": vehicles,
        "daily_routes": daily_routes,
        "selected_date": filter_date.strftime('%Y-%m-%d'),
        "selected_date_display": filter_date.strftime('%d/%m/%Y'),
        "deleted_all": deleted_all
    })

@app.post("/daily-new/add")
async def add_daily_new_route(request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    
    # Lấy ngày được chọn từ form
    selected_date_str = form_data.get("date")
    if not selected_date_str:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = date.today()
    
    # Lấy tất cả routes và sắp xếp theo mã tuyến (A-Z)
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Xử lý từng route
    for route in routes:
        route_id = route.id
        
        # Lấy dữ liệu từ form cho route này
        distance_km = form_data.get(f"distance_km_{route_id}")
        driver_name = form_data.get(f"driver_name_{route_id}")
        license_plate = form_data.get(f"license_plate_{route_id}")
        notes = form_data.get(f"notes_{route_id}")
        
        # Chỉ tạo record nếu có ít nhất một trường được điền
        if distance_km or driver_name or license_plate or notes:
            daily_route = DailyRoute(
                route_id=route_id,
                date=selected_date,
                distance_km=float(distance_km) if distance_km else 0,
                cargo_weight=0,  # Set default value
                driver_name=driver_name or "",
                license_plate=license_plate or "",
                employee_name="",  # Empty since we removed this field
                notes=notes or ""
            )
            db.add(daily_route)
    
    db.commit()
    # Redirect về trang daily-new với ngày đã chọn
    return RedirectResponse(url=f"/daily-new?selected_date={selected_date.strftime('%Y-%m-%d')}", status_code=303)

@app.get("/daily-new/edit/{daily_route_id}", response_class=HTMLResponse)
async def edit_daily_new_route_page(request: Request, daily_route_id: int, db: Session = Depends(get_db)):
    """Trang sửa chuyến"""
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if not daily_route:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Lấy danh sách để hiển thị trong dropdown
    employees = db.query(Employee).filter(Employee.status == 1).all()
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Sắp xếp employees theo tên (A-Z) để dễ tìm kiếm trong dropdown
    employees = sorted(employees, key=lambda emp: emp.name.lower() if emp.name else "")
    
    return templates.TemplateResponse("edit_daily_route.html", {
        "request": request,
        "daily_route": daily_route,
        "employees": employees,
        "vehicles": vehicles
    })

@app.post("/daily-new/edit/{daily_route_id}")
async def edit_daily_new_route(
    daily_route_id: int,
    distance_km: float = Form(0),
    driver_name: str = Form(""),
    license_plate: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    """Cập nhật chuyến"""
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if not daily_route:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Cập nhật thông tin
    daily_route.distance_km = distance_km
    daily_route.driver_name = driver_name
    daily_route.license_plate = license_plate
    daily_route.notes = notes
    
    db.commit()
    
    # Redirect về trang daily-new với ngày của chuyến
    return RedirectResponse(url=f"/daily-new?selected_date={daily_route.date.strftime('%Y-%m-%d')}", status_code=303)

@app.post("/daily-new/delete/{daily_route_id}")
async def delete_daily_new_route(daily_route_id: int, db: Session = Depends(get_db)):
    daily_route = db.query(DailyRoute).filter(DailyRoute.id == daily_route_id).first()
    if daily_route:
        # Lưu ngày của chuyến bị xóa để redirect về đúng ngày
        deleted_date = daily_route.date
        db.delete(daily_route)
        db.commit()
        return RedirectResponse(url=f"/daily-new?selected_date={deleted_date.strftime('%Y-%m-%d')}", status_code=303)
    return RedirectResponse(url="/daily-new", status_code=303)

@app.post("/daily-new/delete-all")
async def delete_all_daily_routes(request: Request, db: Session = Depends(get_db)):
    """Xóa tất cả chuyến đã ghi nhận trong một ngày"""
    form_data = await request.form()
    selected_date_str = form_data.get("date")
    
    if not selected_date_str:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        return RedirectResponse(url="/daily-new", status_code=303)
    
    # Tìm và xóa tất cả chuyến trong ngày được chọn
    daily_routes = db.query(DailyRoute).filter(DailyRoute.date == selected_date).all()
    
    if daily_routes:
        for daily_route in daily_routes:
            db.delete(daily_route)
        db.commit()
    
    # Redirect về trang daily-new với ngày đã chọn và thông báo thành công
    return RedirectResponse(url=f"/daily-new?selected_date={selected_date.strftime('%Y-%m-%d')}&deleted_all=true", status_code=303)

@app.get("/salary/driver-details/{driver_name}")
async def get_driver_details(
    driver_name: str,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Lấy chi tiết chuyến của một lái xe cụ thể"""
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = db.query(DailyRoute).filter(
                DailyRoute.driver_name == driver_name,
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            return {"error": "Invalid date format"}
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        daily_routes_query = db.query(DailyRoute).filter(
            DailyRoute.driver_name == driver_name,
            DailyRoute.date >= date(today.year, today.month, 1),
            DailyRoute.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    # Lấy dữ liệu và join với Route để có thông tin tuyến
    daily_routes = daily_routes_query.join(Route).order_by(DailyRoute.date.desc()).all()
    
    # Format dữ liệu
    trip_details = []
    for trip in daily_routes:
        trip_details.append({
            'date': trip.date.strftime('%d/%m/%Y'),
            'route_code': trip.route.route_code,
            'route_name': trip.route.route_name,
            'license_plate': trip.license_plate,
            'distance_km': trip.distance_km,
            'cargo_weight': trip.cargo_weight,
            'notes': trip.notes or ''
        })
    
    return {"trip_details": trip_details}

@app.get("/salary/driver-details-page/{driver_name}", response_class=HTMLResponse)
async def driver_details_page(
    request: Request,
    driver_name: str,
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Trang hiển thị chi tiết chuyến của một lái xe cụ thể"""
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = db.query(DailyRoute).filter(
                DailyRoute.driver_name == driver_name,
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
            period_text = f"từ {from_date_obj.strftime('%d/%m/%Y')} đến {to_date_obj.strftime('%d/%m/%Y')}"
        except ValueError:
            return RedirectResponse(url="/salary", status_code=303)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        daily_routes_query = db.query(DailyRoute).filter(
            DailyRoute.driver_name == driver_name,
            DailyRoute.date >= date(today.year, today.month, 1),
            DailyRoute.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
        period_text = f"tháng {today.month}/{today.year}"
    
    # Lấy dữ liệu và join với Route để có thông tin tuyến
    daily_routes = daily_routes_query.join(Route).order_by(DailyRoute.date.desc()).all()
    
    # Tính thống kê
    total_trips = len(daily_routes)
    total_distance = sum(trip.distance_km for trip in daily_routes)
    total_cargo = sum(trip.cargo_weight for trip in daily_routes)
    routes_used = list(set(trip.route.route_code for trip in daily_routes))
    
    return templates.TemplateResponse("driver_details.html", {
        "request": request,
        "driver_name": driver_name,
        "period_text": period_text,
        "daily_routes": daily_routes,
        "total_trips": total_trips,
        "total_distance": total_distance,
        "total_cargo": total_cargo,
        "routes_used": routes_used,
        "from_date": from_date,
        "to_date": to_date
    })



@app.get("/salary-simple", response_class=HTMLResponse)
async def salary_simple_page(request: Request):
    """Redirect đến trang báo cáo tổng hợp"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/report", status_code=302)

@app.get("/general-report", response_class=HTMLResponse)
async def general_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Trang thống kê tổng hợp - báo cáo chi tiết hoạt động vận chuyển"""
    # Khởi tạo query cơ bản
    daily_routes_query = db.query(DailyRoute)
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = daily_routes_query.filter(
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc khác
    if driver_name:
        daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name.ilike(f"%{driver_name}%"))
    if license_plate:
        daily_routes_query = daily_routes_query.filter(DailyRoute.license_plate.ilike(f"%{license_plate}%"))
    if route_code:
        daily_routes_query = daily_routes_query.join(Route).filter(Route.route_code.ilike(f"%{route_code}%"))
    
    daily_routes = daily_routes_query.all()
    
    # Tính thống kê theo lái xe
    driver_stats = {}
    for daily_route in daily_routes:
        driver_name_key = daily_route.driver_name
        license_plate_key = daily_route.license_plate
        if driver_name_key and driver_name_key not in driver_stats:
            driver_stats[driver_name_key] = {
                'driver_name': driver_name_key,
                'license_plate': license_plate_key or 'N/A',
                'trip_count': 0,
                'total_distance': 0,
                'total_cargo': 0,
                'routes': set()
            }
        
        if driver_name_key:
            driver_stats[driver_name_key]['trip_count'] += 1
            driver_stats[driver_name_key]['total_distance'] += daily_route.distance_km
            driver_stats[driver_name_key]['total_cargo'] += daily_route.cargo_weight
            driver_stats[driver_name_key]['routes'].add(daily_route.route.route_code)
            if license_plate_key:
                driver_stats[driver_name_key]['license_plate'] = license_plate_key
    
    # Convert to list
    salary_data = []
    for driver_name_key, stats in driver_stats.items():
        salary_data.append({
            'driver_name': driver_name_key,
            'license_plate': stats['license_plate'],
            'trip_count': stats['trip_count'],
            'total_distance': stats['total_distance'],
            'total_cargo': stats['total_cargo'],
            'routes': list(stats['routes'])
        })
    
    salary_data.sort(key=lambda x: x['trip_count'], reverse=True)
    
    # Tạo dữ liệu chi tiết từng chuyến
    trip_details = []
    for daily_route in daily_routes:
        if daily_route.driver_name:
            trip_details.append({
                'driver_name': daily_route.driver_name,
                'license_plate': daily_route.license_plate or 'N/A',
                'date': daily_route.date,
                'route_code': daily_route.route.route_code,
                'route_name': daily_route.route.route_name,
                'distance_km': daily_route.distance_km,
                'cargo_weight': daily_route.cargo_weight,
                'notes': daily_route.notes or ''
            })
    
    trip_details.sort(key=lambda x: (x['driver_name'], x['date']))
    
    # Lấy danh sách cho dropdown
    routes = db.query(Route).all()
    employees = db.query(Employee).all()
    vehicles = db.query(Vehicle).all()
    
    # Template data - CHỈ TRUYỀN KHI CÓ GIÁ TRỊ
    template_data = {
        "request": request,
        "salary_data": salary_data,
        "trip_details": trip_details,
        "employees": employees,
        "vehicles": vehicles,
        "routes": routes,
        "total_routes": len(daily_routes),
        "total_distance": sum(dr.distance_km for dr in daily_routes),
        "total_cargo": sum(dr.cargo_weight for dr in daily_routes)
    }
    
    # Chỉ thêm khi có giá trị
    if from_date:
        template_data["from_date"] = from_date
    if to_date:
        template_data["to_date"] = to_date
    if driver_name:
        template_data["driver_name"] = driver_name
    if license_plate:
        template_data["license_plate"] = license_plate
    if route_code:
        template_data["route_code"] = route_code
    
    return templates.TemplateResponse("salary_simple.html", template_data)

@app.get("/salary-simple/export-excel")
async def export_salary_simple_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Redirect đến general-report export"""
    from fastapi.responses import RedirectResponse
    params = []
    if from_date:
        params.append(f"from_date={from_date}")
    if to_date:
        params.append(f"to_date={to_date}")
    if driver_name:
        params.append(f"driver_name={driver_name}")
    if license_plate:
        params.append(f"license_plate={license_plate}")
    if route_code:
        params.append(f"route_code={route_code}")
    
    url = "/general-report/export-excel"
    if params:
        url += "?" + "&".join(params)
    
    return RedirectResponse(url=url, status_code=302)

@app.get("/general-report/export-excel")
async def export_general_report_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    driver_name: Optional[str] = None,
    license_plate: Optional[str] = None,
    route_code: Optional[str] = None
):
    """Xuất Excel danh sách chi tiết từng chuyến cho general-report"""
    # Sử dụng lại logic lọc từ salary_simple_page
    daily_routes_query = db.query(DailyRoute)
    
    # Áp dụng bộ lọc thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            daily_routes_query = daily_routes_query.filter(
                DailyRoute.date >= from_date_obj,
                DailyRoute.date <= to_date_obj
            )
        except ValueError:
            pass
    
    # Áp dụng các bộ lọc khác
    if driver_name:
        daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name.ilike(f"%{driver_name}%"))
    if license_plate:
        daily_routes_query = daily_routes_query.filter(DailyRoute.license_plate.ilike(f"%{license_plate}%"))
    if route_code:
        daily_routes_query = daily_routes_query.join(Route).filter(Route.route_code.ilike(f"%{route_code}%"))
    
    daily_routes = daily_routes_query.all()
    
    # Tạo dữ liệu chi tiết từng chuyến
    trip_details = []
    for daily_route in daily_routes:
        if daily_route.driver_name:
            trip_details.append({
                'stt': len(trip_details) + 1,
                'ngay_chay': daily_route.date.strftime('%d/%m/%Y'),
                'ten_lai_xe': daily_route.driver_name,
                'bien_so_xe': daily_route.license_plate or 'N/A',
                'ma_tuyen': daily_route.route.route_code,
                'ten_tuyen': daily_route.route.route_name,
                'km': daily_route.distance_km,
                'tai_trong': daily_route.cargo_weight,
                'ghi_chu': daily_route.notes or ''
            })
    
    # Tạo CSV content với UTF-8 BOM để Excel hiển thị đúng tiếng Việt
    csv_content = "\ufeff"  # UTF-8 BOM
    csv_content += "STT,Ngày chạy,Tên lái xe,Biển số xe,Mã tuyến,Tên tuyến,Km,Tải trọng,Ghi chú\n"
    
    for trip in trip_details:
        # Escape các ký tự đặc biệt trong CSV
        def escape_csv_field(field):
            if field is None:
                return ""
            field_str = str(field)
            # Nếu chứa dấu phẩy, dấu ngoặc kép hoặc xuống dòng thì bọc trong dấu ngoặc kép
            if ',' in field_str or '"' in field_str or '\n' in field_str:
                field_str = field_str.replace('"', '""')  # Escape dấu ngoặc kép
                field_str = f'"{field_str}"'
            return field_str
        
        csv_content += f"{trip['stt']},{escape_csv_field(trip['ngay_chay'])},{escape_csv_field(trip['ten_lai_xe'])},{escape_csv_field(trip['bien_so_xe'])},{escape_csv_field(trip['ma_tuyen'])},{escape_csv_field(trip['ten_tuyen'])},{trip['km']},{trip['tai_trong']},{escape_csv_field(trip['ghi_chu'])}\n"
    
    # Tạo tên file
    if from_date and to_date:
        filename = f"chi_tiet_chuyen_{from_date}_den_{to_date}.csv"
    else:
        today = date.today()
        filename = f"chi_tiet_chuyen_{today.month}_{today.year}.csv"
    
    # Trả về file CSV với encoding UTF-8
    return Response(
        content=csv_content.encode('utf-8-sig'),  # UTF-8 with BOM
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Content-Type": "text/csv; charset=utf-8"
        }
    )



# ===== FUEL MANAGEMENT ROUTES =====

@app.get("/fuel", response_class=HTMLResponse)
async def fuel_page(request: Request):
    """Redirect đến trang báo cáo tổng hợp"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/report", status_code=302)

@app.get("/fuel-report", response_class=HTMLResponse)
async def fuel_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Trang tổng hợp đổ dầu - báo cáo chi tiết"""
    # Xử lý khoảng thời gian
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            fuel_records_query = db.query(FuelRecord).filter(
                FuelRecord.date >= from_date_obj,
                FuelRecord.date <= to_date_obj
            )
        except ValueError:
            fuel_records_query = db.query(FuelRecord)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        fuel_records_query = db.query(FuelRecord).filter(
            FuelRecord.date >= date(today.year, today.month, 1),
            FuelRecord.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.license_plate).all()
    
    # Tính tổng số lít dầu đã đổ
    total_liters_pumped = sum(record.liters_pumped for record in fuel_records)
    
    # Lấy danh sách xe để hiển thị trong dropdown
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    # Tạo template data
    template_data = {
        "request": request,
        "fuel_records": fuel_records,
        "vehicles": vehicles,
        "total_liters_pumped": total_liters_pumped,
        "total_records": len(fuel_records)
    }
    
    if from_date:
        template_data["from_date"] = from_date
    if to_date:
        template_data["to_date"] = to_date
    
    return templates.TemplateResponse("fuel.html", template_data)

@app.post("/fuel/add")
async def add_fuel_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Thêm bản ghi đổ dầu mới"""
    form_data = await request.form()
    
    # Lấy dữ liệu từ form
    date_str = form_data.get("date")
    fuel_type = form_data.get("fuel_type", "Dầu DO 0,05S-II")
    license_plate = form_data.get("license_plate")
    fuel_price_per_liter = float(form_data.get("fuel_price_per_liter", 0))
    liters_pumped = float(form_data.get("liters_pumped", 0))
    notes = form_data.get("notes", "")
    
    if not date_str or not license_plate:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    try:
        fuel_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        fuel_date = date.today()
    
    # Tính toán số tiền dầu đã đổ = Đơn giá dầu × Số lít dầu đã đổ (làm tròn đến đồng)
    cost_pumped = round(fuel_price_per_liter * liters_pumped)
    
    # Tạo bản ghi mới
    fuel_record = FuelRecord(
        date=fuel_date,
        fuel_type=fuel_type,
        license_plate=license_plate,
        fuel_price_per_liter=fuel_price_per_liter,
        liters_pumped=liters_pumped,
        cost_pumped=cost_pumped,
        notes=notes
    )
    
    db.add(fuel_record)
    db.commit()
    
    # Redirect với tham số thời gian nếu có
    redirect_url = "/fuel-report"
    from_date = form_data.get("from_date")
    to_date = form_data.get("to_date")
    if from_date and to_date:
        redirect_url += f"?from_date={from_date}&to_date={to_date}"
    
    return RedirectResponse(url=redirect_url, status_code=303)

@app.post("/fuel/delete/{fuel_record_id}")
async def delete_fuel_record(
    fuel_record_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Xóa bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if fuel_record:
        db.delete(fuel_record)
        db.commit()
    
    # Redirect về trang fuel
    return RedirectResponse(url="/fuel-report", status_code=303)

@app.get("/fuel/edit/{fuel_record_id}", response_class=HTMLResponse)
async def edit_fuel_record_page(
    request: Request,
    fuel_record_id: int,
    db: Session = Depends(get_db)
):
    """Trang sửa bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if not fuel_record:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    
    return templates.TemplateResponse("edit_fuel.html", {
        "request": request,
        "fuel_record": fuel_record,
        "vehicles": vehicles
    })

@app.post("/fuel/edit/{fuel_record_id}")
async def edit_fuel_record(
    fuel_record_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Cập nhật bản ghi đổ dầu"""
    fuel_record = db.query(FuelRecord).filter(FuelRecord.id == fuel_record_id).first()
    if not fuel_record:
        return RedirectResponse(url="/fuel-report", status_code=303)
    
    form_data = await request.form()
    
    # Cập nhật dữ liệu
    date_str = form_data.get("date")
    if date_str:
        try:
            fuel_record.date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    fuel_record.fuel_type = form_data.get("fuel_type", "Dầu DO 0,05S-II")
    fuel_record.license_plate = form_data.get("license_plate")
    fuel_record.fuel_price_per_liter = float(form_data.get("fuel_price_per_liter", 0))
    fuel_record.liters_pumped = float(form_data.get("liters_pumped", 0))
    fuel_record.notes = form_data.get("notes", "")
    
    # Tính toán lại số tiền dầu đã đổ = Đơn giá dầu × Số lít dầu đã đổ (làm tròn đến đồng)
    fuel_record.cost_pumped = round(fuel_record.fuel_price_per_liter * fuel_record.liters_pumped)
    
    db.commit()
    return RedirectResponse(url="/fuel-report", status_code=303)

@app.get("/fuel/download-template")
async def download_fuel_template(db: Session = Depends(get_db)):
    """Tải mẫu Excel để import dữ liệu đổ dầu"""
    # Lấy danh sách xe để hiển thị trong mẫu
    vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
    vehicle_list = [v.license_plate for v in vehicles]
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mẫu Import Đổ Dầu"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề
    ws.merge_cells('A1:F1')
    ws['A1'] = "MẪU IMPORT DỮ LIỆU ĐỔ DẦU"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Hướng dẫn
    ws.merge_cells('A2:F2')
    ws['A2'] = "Vui lòng điền dữ liệu theo đúng định dạng bên dưới"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Header bảng
    headers = [
        "STT", "Ngày đổ dầu (dd/mm/yyyy)", "Biển số xe", 
        "Số lượng dầu đổ (lít)", "Đơn giá (đồng/lít)", "Thành tiền (đồng)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu mẫu
    sample_data = [
        [1, "01/01/2025", "51A-12345", 50.000, 19020, 951000],
        [2, "02/01/2025", "51B-67890", 45.500, 19020, 865410],
        [3, "03/01/2025", "51C-11111", 60.000, 19020, 1141200]
    ]
    
    for row, data in enumerate(sample_data, 5):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Định dạng số
    for row in range(5, 8):
        # Số lượng dầu - 3 chữ số thập phân
        ws.cell(row=row, column=4).number_format = '#,##0.000'
        # Đơn giá - 2 chữ số thập phân
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # Thành tiền - không có chữ số thập phân
        ws.cell(row=row, column=6).number_format = '#,##0'
    
    # Thêm sheet hướng dẫn
    ws2 = wb.create_sheet("Hướng dẫn")
    ws2['A1'] = "HƯỚNG DẪN SỬ DỤNG"
    ws2['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "1. Định dạng cột:",
        "   - STT: Số thứ tự (tự động)",
        "   - Ngày đổ dầu: Định dạng dd/mm/yyyy (ví dụ: 01/01/2025)",
        "   - Biển số xe: Phải khớp với danh sách xe trong hệ thống",
        "   - Số lượng dầu đổ: Cho phép 3 chữ số thập phân (ví dụ: 50.000)",
        "   - Đơn giá: Số chính xác (ví dụ: 19020)",
        "   - Thành tiền: Có thể để trống, hệ thống sẽ tự tính",
        "",
        "2. Danh sách biển số xe hợp lệ:",
    ]
    
    for i, instruction in enumerate(instructions, 2):
        ws2.cell(row=i, column=1, value=instruction)
    
    # Thêm danh sách xe
    for i, vehicle in enumerate(vehicle_list, len(instructions) + 2):
        ws2.cell(row=i, column=1, value=f"   - {vehicle}")
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 20, 15, 20, 20, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws2.column_dimensions['A'].width = 50
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    today = date.today()
    filename = f"Mau_Import_DoDau_{today.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.post("/fuel/import-excel")
async def import_fuel_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Import dữ liệu đổ dầu từ file Excel"""
    try:
        # Kiểm tra định dạng file
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return JSONResponse(
                status_code=400,
                content={
                    "success": False, 
                    "error": "Định dạng file không hợp lệ",
                    "error_type": "file_format",
                    "details": "Chỉ chấp nhận file Excel (.xlsx hoặc .xls)",
                    "suggestion": "Vui lòng chọn file Excel có định dạng .xlsx hoặc .xls"
                }
            )
        
        # Đọc file Excel
        content = await file.read()
        if len(content) == 0:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "File rỗng",
                    "error_type": "empty_file",
                    "details": "File Excel không chứa dữ liệu",
                    "suggestion": "Vui lòng kiểm tra lại file Excel có chứa dữ liệu"
                }
            )
        
        try:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
        except Exception as e:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "error": "Không thể đọc file Excel",
                    "error_type": "file_corrupted",
                    "details": f"Lỗi kỹ thuật: {str(e)}",
                    "suggestion": "Vui lòng kiểm tra file Excel không bị hỏng và có định dạng đúng"
                }
            )
        
        # Lấy danh sách xe hợp lệ
        vehicles = db.query(Vehicle).filter(Vehicle.status == 1).all()
        valid_license_plates = {v.license_plate for v in vehicles}
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Bỏ qua header (dòng 1-4)
        for row_num in range(5, ws.max_row + 1):
            try:
                # Đọc dữ liệu từ Excel
                stt = ws.cell(row=row_num, column=1).value
                date_str = ws.cell(row=row_num, column=2).value
                license_plate = ws.cell(row=row_num, column=3).value
                liters_pumped = ws.cell(row=row_num, column=4).value
                fuel_price_per_liter = ws.cell(row=row_num, column=5).value
                cost_pumped = ws.cell(row=row_num, column=6).value
                
                # Bỏ qua dòng trống
                if not date_str or not license_plate:
                    continue
                
                # Validation dữ liệu với thông báo chi tiết
                validation_errors = []
                
                # Kiểm tra ngày (cột B)
                if isinstance(date_str, str):
                    try:
                        fuel_date = datetime.strptime(date_str, "%d/%m/%Y").date()
                    except ValueError:
                        validation_errors.append({
                            "column": "B (Ngày đổ)",
                            "error": "Định dạng ngày không đúng",
                            "value": str(date_str),
                            "suggestion": "Định dạng đúng: dd/mm/yyyy (ví dụ: 25/09/2025)"
                        })
                elif isinstance(date_str, datetime):
                    fuel_date = date_str.date()
                else:
                    validation_errors.append({
                        "column": "B (Ngày đổ)",
                        "error": "Ngày không hợp lệ",
                        "value": str(date_str),
                        "suggestion": "Vui lòng nhập ngày theo định dạng dd/mm/yyyy"
                    })
                
                # Kiểm tra biển số xe (cột C)
                if not license_plate:
                    validation_errors.append({
                        "column": "C (Biển số xe)",
                        "error": "Biển số xe không được để trống",
                        "value": "",
                        "suggestion": "Vui lòng nhập biển số xe"
                    })
                elif str(license_plate).strip() not in valid_license_plates:
                    validation_errors.append({
                        "column": "C (Biển số xe)",
                        "error": "Biển số xe không tồn tại trong hệ thống",
                        "value": str(license_plate),
                        "suggestion": f"Biển số xe hợp lệ: {', '.join(list(valid_license_plates)[:5])}{'...' if len(valid_license_plates) > 5 else ''}"
                    })
                
                # Kiểm tra số lít dầu (cột D)
                try:
                    liters_pumped = float(liters_pumped) if liters_pumped is not None else 0
                    if liters_pumped <= 0:
                        validation_errors.append({
                            "column": "D (Số lít đã đổ)",
                            "error": "Số lít dầu phải lớn hơn 0",
                            "value": str(liters_pumped),
                            "suggestion": "Vui lòng nhập số lít dầu lớn hơn 0 (ví dụ: 50.5)"
                        })
                except (ValueError, TypeError):
                    validation_errors.append({
                        "column": "D (Số lít đã đổ)",
                        "error": "Số lít dầu không hợp lệ",
                        "value": str(liters_pumped),
                        "suggestion": "Vui lòng nhập số lít dầu là số (ví dụ: 50.5, 100)"
                    })
                
                # Kiểm tra đơn giá (cột E)
                try:
                    fuel_price_per_liter = float(fuel_price_per_liter) if fuel_price_per_liter is not None else 0
                    if fuel_price_per_liter <= 0:
                        validation_errors.append({
                            "column": "E (Giá xăng dầu)",
                            "error": "Đơn giá phải lớn hơn 0",
                            "value": str(fuel_price_per_liter),
                            "suggestion": "Vui lòng nhập đơn giá lớn hơn 0 (ví dụ: 25000)"
                        })
                except (ValueError, TypeError):
                    validation_errors.append({
                        "column": "E (Giá xăng dầu)",
                        "error": "Đơn giá không hợp lệ",
                        "value": str(fuel_price_per_liter),
                        "suggestion": "Vui lòng nhập đơn giá là số (ví dụ: 25000, 25000.5)"
                    })
                
                # Tính thành tiền nếu không có
                if cost_pumped is None or cost_pumped == "":
                    cost_pumped = round(fuel_price_per_liter * liters_pumped)
                else:
                    try:
                        cost_pumped = float(cost_pumped)
                    except (ValueError, TypeError):
                        cost_pumped = round(fuel_price_per_liter * liters_pumped)
                
                # Nếu có lỗi validation, bỏ qua dòng này
                if validation_errors:
                    errors.append({
                        "row": row_num,
                        "errors": validation_errors
                    })
                    skipped_count += 1
                    continue
                
                # Kiểm tra trùng lặp (cùng ngày, cùng xe)
                existing_record = db.query(FuelRecord).filter(
                    FuelRecord.date == fuel_date,
                    FuelRecord.license_plate == str(license_plate).strip()
                ).first()
                
                if existing_record:
                    errors.append({
                        "row": row_num,
                        "errors": [{
                            "column": "Tổng hợp",
                            "error": "Bản ghi trùng lặp",
                            "value": f"Xe {license_plate} - Ngày {fuel_date.strftime('%d/%m/%Y')}",
                            "suggestion": "Đã tồn tại bản ghi đổ dầu cho xe này vào ngày này. Vui lòng kiểm tra lại dữ liệu."
                        }]
                    })
                    skipped_count += 1
                    continue
                
                # Tạo bản ghi mới
                fuel_record = FuelRecord(
                    date=fuel_date,
                    fuel_type="Dầu DO 0,05S-II",  # Mặc định
                    license_plate=str(license_plate).strip(),
                    fuel_price_per_liter=fuel_price_per_liter,
                    liters_pumped=liters_pumped,
                    cost_pumped=cost_pumped,
                    notes=f"Import từ Excel - dòng {row_num}"
                )
                
                db.add(fuel_record)
                imported_count += 1
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "errors": [{
                        "column": "Tổng hợp",
                        "error": "Lỗi xử lý dữ liệu",
                        "value": f"Lỗi kỹ thuật: {str(e)}",
                        "suggestion": "Vui lòng kiểm tra định dạng dữ liệu trong dòng này"
                    }]
                })
                skipped_count += 1
                continue
        
        # Commit tất cả thay đổi
        db.commit()
        
        # Tạo response chi tiết
        response_data = {
            "success": True,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "total_errors": len(errors),
            "summary": {
                "total_rows_processed": ws.max_row - 4,  # Trừ header
                "successful_imports": imported_count,
                "failed_imports": skipped_count,
                "success_rate": f"{(imported_count / max(1, ws.max_row - 4)) * 100:.1f}%" if ws.max_row > 4 else "0%"
            }
        }
        
        if errors:
            response_data["errors"] = errors[:20]  # Hiển thị 20 lỗi đầu tiên
            if len(errors) > 20:
                response_data["has_more_errors"] = True
                response_data["remaining_errors"] = len(errors) - 20
            response_data["error_summary"] = {
                "validation_errors": len([e for e in errors if any(err.get("column") != "Tổng hợp" for err in e.get("errors", []))]),
                "duplicate_errors": len([e for e in errors if any("trùng lặp" in err.get("error", "") for err in e.get("errors", []))]),
                "technical_errors": len([e for e in errors if any("Lỗi xử lý" in err.get("error", "") for err in e.get("errors", []))])
            }
        
        return JSONResponse(content=response_data)
        
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=500,
            content={
                "success": False, 
                "error": "Lỗi hệ thống",
                "error_type": "system_error",
                "details": f"Lỗi kỹ thuật: {str(e)}",
                "suggestion": "Vui lòng thử lại hoặc liên hệ quản trị viên nếu lỗi vẫn tiếp tục"
            }
        )

@app.get("/fuel/export-excel")
async def export_fuel_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Redirect đến fuel-report export"""
    from fastapi.responses import RedirectResponse
    params = []
    if from_date:
        params.append(f"from_date={from_date}")
    if to_date:
        params.append(f"to_date={to_date}")
    
    url = "/fuel-report/export-excel"
    if params:
        url += "?" + "&".join(params)
    
    return RedirectResponse(url=url, status_code=302)

@app.get("/fuel-report/export-excel")
async def export_fuel_report_excel(
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Xuất Excel báo cáo đổ dầu"""
    # Xử lý khoảng thời gian (sử dụng logic giống như fuel_page)
    if from_date and to_date:
        try:
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            fuel_records_query = db.query(FuelRecord).filter(
                FuelRecord.date >= from_date_obj,
                FuelRecord.date <= to_date_obj
            )
        except ValueError:
            fuel_records_query = db.query(FuelRecord)
    else:
        # Nếu không có khoảng thời gian, lấy tháng hiện tại
        today = date.today()
        fuel_records_query = db.query(FuelRecord).filter(
            FuelRecord.date >= date(today.year, today.month, 1),
            FuelRecord.date < date(today.year, today.month + 1, 1) if today.month < 12 else date(today.year + 1, 1, 1)
        )
    
    fuel_records = fuel_records_query.order_by(FuelRecord.date.desc(), FuelRecord.license_plate).all()
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo đổ dầu"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:H1')
    ws['A1'] = "BÁO CÁO ĐỔ DẦU"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin thời gian
    period_text = ""
    if from_date and to_date:
        period_text = f"Từ ngày: {from_date} đến ngày: {to_date}"
    else:
        today = date.today()
        period_text = f"Tháng: {today.month}/{today.year}"
    
    ws.merge_cells('A2:H2')
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Header bảng
    headers = [
        "STT", "Ngày đổ", "Loại dầu", "Biển số xe", 
        "Giá xăng dầu (đồng/lít)", "Số lít đã đổ", "Số tiền đã đổ (VNĐ)", "Ghi chú"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    for row, record in enumerate(fuel_records, 5):
        ws.cell(row=row, column=1, value=row-4)  # STT
        ws.cell(row=row, column=2, value=record.date.strftime('%d/%m/%Y'))  # Ngày đổ
        ws.cell(row=row, column=3, value=record.fuel_type)  # Loại dầu
        ws.cell(row=row, column=4, value=record.license_plate)  # Biển số xe
        ws.cell(row=row, column=5, value=record.fuel_price_per_liter)  # Giá xăng dầu
        ws.cell(row=row, column=6, value=record.liters_pumped)  # Số lít đã đổ
        ws.cell(row=row, column=7, value=record.cost_pumped)  # Số tiền đã đổ
        ws.cell(row=row, column=8, value=record.notes or '')  # Ghi chú
    
    # Định dạng số
    for row in range(5, 5 + len(fuel_records)):
        # Giá xăng dầu - 2 chữ số thập phân
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        # Số lít - 3 chữ số thập phân
        ws.cell(row=row, column=6).number_format = '#,##0.000'
        # Số tiền - không có chữ số thập phân
        ws.cell(row=row, column=7).number_format = '#,##0'
    
    # Dòng tổng cộng
    if fuel_records:
        total_row = 5 + len(fuel_records)
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(r.liters_pumped for r in fuel_records)).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=sum(r.cost_pumped for r in fuel_records)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value="").font = Font(bold=True)
        
        # Định dạng số cho dòng tổng cộng
        ws.cell(row=total_row, column=6).number_format = '#,##0.000'
        ws.cell(row=total_row, column=7).number_format = '#,##0'
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 12, 20, 15, 20, 15, 18, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    today = date.today()
    filename = f"BaoCao_DoDau_{today.strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

# ===== SALARY CALCULATION ROUTES =====

@app.get("/api/employees")
async def get_employees_api(db: Session = Depends(get_db)):
    """API để lấy danh sách nhân viên cho dropdown"""
    employees = db.query(Employee).filter(Employee.status == 1).all()
    return [
        {
            "id": emp.id,
            "name": emp.name
        }
        for emp in employees
    ]

@app.get("/salary-calculation", response_class=HTMLResponse)
async def salary_calculation_page(
    request: Request, 
    db: Session = Depends(get_db),
    selected_month: Optional[str] = None,
    selected_employee: Optional[str] = None,
    selected_route: Optional[str] = None
):
    """Trang bảng tính lương"""
    import calendar
    
    # Xử lý tháng được chọn
    if selected_month:
        try:
            # selected_month format: "2025-01"
            year, month = selected_month.split('-')
            year, month = int(year), int(month)
        except ValueError:
            # Nếu format không đúng, dùng tháng hiện tại
            today = date.today()
            year, month = today.year, today.month
    else:
        # Nếu không có tháng được chọn, dùng tháng hiện tại
        today = date.today()
        year, month = today.year, today.month
    
    # Tính số ngày trong tháng
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Lấy dữ liệu chuyến trong tháng được chọn
    from_date = date(year, month, 1)
    to_date = date(year, month, days_in_month)
    
    # Xây dựng query cơ bản
    daily_routes_query = db.query(DailyRoute).filter(
        DailyRoute.date >= from_date,
        DailyRoute.date <= to_date,
        DailyRoute.driver_name.isnot(None),
        DailyRoute.driver_name != ""
    )
    
    # Thêm filter theo nhân viên nếu được chọn
    if selected_employee and selected_employee != "all":
        # Tìm nhân viên theo ID hoặc tên
        try:
            employee_id = int(selected_employee)
            employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
            if employee:
                daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == employee.name)
        except ValueError:
            # Nếu không phải số, coi như tên nhân viên
            daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == selected_employee)
    
    # Join với Route để có thể filter theo route_code
    daily_routes_query = daily_routes_query.join(Route)
    
    # Thêm filter theo mã tuyến nếu được chọn
    if selected_route and selected_route != "all":
        daily_routes_query = daily_routes_query.filter(Route.route_code == selected_route)
    
    daily_routes = daily_routes_query.order_by(Route.route_code, DailyRoute.date).all()
    
    # Tính lương cho từng chuyến và lấy biển số xe
    salary_data = []
    for daily_route in daily_routes:
        # Tính lương theo công thức khác nhau tùy loại tuyến
        daily_salary = 0
        salary_type = "standard"  # Mặc định là tuyến chuẩn
        
        # Kiểm tra nếu là tuyến "Tăng Cường"
        if daily_route.route.route_code and daily_route.route.route_code.strip() == "Tăng Cường":
            salary_type = "tang_cuong"  # Luôn đánh dấu là tuyến Tăng Cường
            # Công thức cho tuyến "Tăng Cường": Số km thực tế × 1,100 đ
            if daily_route.distance_km and daily_route.distance_km > 0:
                daily_salary = daily_route.distance_km * 1100
        else:
            # Công thức cho tuyến thường: Lương tuyến/tháng / 30
            if daily_route.route.monthly_salary and daily_route.route.monthly_salary > 0:
                daily_salary = daily_route.route.monthly_salary / 30
        
        # Lấy biển số xe từ daily-new với điều kiện lọc chính xác:
        # Tên nhân viên + Mã tuyến + Ngày chạy
        license_plate_display = "Chưa cập nhật"
        if daily_route.driver_name:
            # Tìm chuyến có cùng: tên lái xe + route_id + ngày chạy
            matching_routes = db.query(DailyRoute).filter(
                DailyRoute.driver_name == daily_route.driver_name,
                DailyRoute.route_id == daily_route.route_id,
                DailyRoute.date == daily_route.date,
                DailyRoute.license_plate.isnot(None),
                DailyRoute.license_plate != ""
            ).order_by(DailyRoute.created_at.desc()).all()  # Sắp xếp theo thời gian tạo mới nhất
            
            if matching_routes:
                # Lấy danh sách biển số xe duy nhất từ các chuyến khớp
                license_plates = list(set([route.license_plate for route in matching_routes if route.license_plate]))
                
                if license_plates:
                    if len(license_plates) == 1:
                        license_plate_display = license_plates[0]
                    else:
                        # Nếu có nhiều biển số, hiển thị phân tách bằng dấu phẩy
                        license_plate_display = ", ".join(license_plates)
        
        salary_data.append({
            'driver_name': daily_route.driver_name,
            'route_code': daily_route.route.route_code,
            'route_name': daily_route.route.route_name,
            'date': daily_route.date,
            'license_plate': license_plate_display,
            'daily_salary': daily_salary,
            'monthly_salary': daily_route.route.monthly_salary or 0,
            'days_in_month': 30,  # Chuẩn hóa tháng 30 ngày
            'salary_type': salary_type,  # "standard" hoặc "tang_cuong"
            'distance_km': daily_route.distance_km or 0  # Số km thực tế cho tuyến Tăng Cường
        })
    
    # Lấy danh sách lái xe và tuyến để hiển thị
    employees = db.query(Employee).filter(Employee.status == 1).all()
    routes = db.query(Route).filter(Route.is_active == 1, Route.status == 1).all()
    
    # Sắp xếp routes: A-Z bình thường, nhưng "Tăng Cường" đẩy xuống cuối
    def sort_routes_with_tang_cuong_at_bottom(routes):
        # Lọc ra routes không phải "Tăng Cường"
        normal_routes = [route for route in routes if route.route_code and route.route_code.strip() != "Tăng Cường"]
        
        # Lọc ra routes "Tăng Cường"
        tang_cuong_routes = [route for route in routes if route.route_code and route.route_code.strip() == "Tăng Cường"]
        
        # Sắp xếp routes bình thường theo A-Z
        normal_routes_sorted = sorted(normal_routes, key=lambda route: route.route_code.lower())
        
        # Ghép lại: routes bình thường + routes "Tăng Cường"
        return normal_routes_sorted + tang_cuong_routes
    
    routes = sort_routes_with_tang_cuong_at_bottom(routes)
    
    # Tính tổng lương theo loại tuyến
    total_standard_salary = sum(item['daily_salary'] for item in salary_data if item['salary_type'] == 'standard')
    total_tang_cuong_salary = sum(item['daily_salary'] for item in salary_data if item['salary_type'] == 'tang_cuong')
    total_salary = total_standard_salary + total_tang_cuong_salary
    
    # Tạo template data
    template_data = {
        "request": request,
        "salary_data": salary_data,
        "employees": employees,
        "routes": routes,
        "selected_month": f"{year}-{month:02d}",
        "selected_month_display": f"{month}/{year}",
        "selected_employee": selected_employee or "all",
        "selected_route": selected_route or "all",
        "days_in_month": days_in_month,
        "total_trips": len(salary_data),
        "total_salary": total_salary,
        "total_standard_salary": total_standard_salary,
        "total_tang_cuong_salary": total_tang_cuong_salary
    }
    
    return templates.TemplateResponse("salary_calculation.html", template_data)

@app.get("/salary-calculation/export-excel")
async def export_salary_calculation_excel(
    db: Session = Depends(get_db),
    selected_month: Optional[str] = None,
    selected_employee: Optional[str] = None,
    selected_route: Optional[str] = None
):
    """Xuất Excel bảng tính lương"""
    import calendar
    
    # Xử lý tháng được chọn (sử dụng logic giống như salary_calculation_page)
    if selected_month:
        try:
            year, month = selected_month.split('-')
            year, month = int(year), int(month)
        except ValueError:
            today = date.today()
            year, month = today.year, today.month
    else:
        today = date.today()
        year, month = today.year, today.month
    
    # Tính số ngày trong tháng
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Lấy dữ liệu chuyến trong tháng được chọn
    from_date = date(year, month, 1)
    to_date = date(year, month, days_in_month)
    
    # Xây dựng query cơ bản (sử dụng logic giống như salary_calculation_page)
    daily_routes_query = db.query(DailyRoute).filter(
        DailyRoute.date >= from_date,
        DailyRoute.date <= to_date,
        DailyRoute.driver_name.isnot(None),
        DailyRoute.driver_name != ""
    )
    
    # Thêm filter theo nhân viên nếu được chọn
    if selected_employee and selected_employee != "all":
        try:
            employee_id = int(selected_employee)
            employee = db.query(Employee).filter(Employee.id == employee_id, Employee.status == 1).first()
            if employee:
                daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == employee.name)
        except ValueError:
            daily_routes_query = daily_routes_query.filter(DailyRoute.driver_name == selected_employee)
    
    # Join với Route để có thể filter theo route_code
    daily_routes_query = daily_routes_query.join(Route)
    
    # Thêm filter theo mã tuyến nếu được chọn
    if selected_route and selected_route != "all":
        daily_routes_query = daily_routes_query.filter(Route.route_code == selected_route)
    
    daily_routes = daily_routes_query.order_by(Route.route_code, DailyRoute.date).all()
    
    # Tính lương cho từng chuyến và lấy biển số xe (sử dụng logic giống như salary_calculation_page)
    salary_data = []
    for daily_route in daily_routes:
        # Tính lương theo công thức khác nhau tùy loại tuyến
        daily_salary = 0
        salary_type = "standard"  # Mặc định là tuyến chuẩn
        
        # Kiểm tra nếu là tuyến "Tăng Cường"
        if daily_route.route.route_code and daily_route.route.route_code.strip() == "Tăng Cường":
            salary_type = "tang_cuong"  # Luôn đánh dấu là tuyến Tăng Cường
            # Công thức cho tuyến "Tăng Cường": Số km thực tế × 1,100 đ
            if daily_route.distance_km and daily_route.distance_km > 0:
                daily_salary = daily_route.distance_km * 1100
        else:
            # Công thức cho tuyến thường: Lương tuyến/tháng / 30
            if daily_route.route.monthly_salary and daily_route.route.monthly_salary > 0:
                daily_salary = daily_route.route.monthly_salary / 30
        
        # Lấy biển số xe từ daily-new với điều kiện lọc chính xác
        license_plate_display = "Chưa cập nhật"
        if daily_route.driver_name:
            matching_routes = db.query(DailyRoute).filter(
                DailyRoute.driver_name == daily_route.driver_name,
                DailyRoute.route_id == daily_route.route_id,
                DailyRoute.date == daily_route.date,
                DailyRoute.license_plate.isnot(None),
                DailyRoute.license_plate != ""
            ).order_by(DailyRoute.created_at.desc()).all()
            
            if matching_routes:
                license_plates = list(set([route.license_plate for route in matching_routes if route.license_plate]))
                
                if license_plates:
                    if len(license_plates) == 1:
                        license_plate_display = license_plates[0]
                    else:
                        license_plate_display = ", ".join(license_plates)
        
        salary_data.append({
            'driver_name': daily_route.driver_name,
            'route_code': daily_route.route.route_code,
            'route_name': daily_route.route.route_name,
            'date': daily_route.date,
            'license_plate': license_plate_display,
            'daily_salary': daily_salary,
            'salary_type': salary_type,  # "standard" hoặc "tang_cuong"
            'distance_km': daily_route.distance_km or 0  # Số km thực tế cho tuyến Tăng Cường
        })
    
    # Tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng tính lương"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:F1')
    ws['A1'] = "BẢNG TÍNH LƯƠNG"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thông tin tháng
    month_text = f"Tháng: {month}/{year}"
    ws.merge_cells('A2:F2')
    ws['A2'] = month_text
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(italic=True)
    
    # Header bảng
    headers = [
        "STT", "Họ và tên lái xe", "Mã tuyến", 
        "Ngày chạy", "Biển số xe", "Lương chuyến (VNĐ)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    for row, item in enumerate(salary_data, 5):
        ws.cell(row=row, column=1, value=row-4)  # STT
        ws.cell(row=row, column=2, value=item['driver_name'])  # Họ và tên lái xe
        ws.cell(row=row, column=3, value=item['route_code'])  # Mã tuyến
        ws.cell(row=row, column=4, value=item['date'].strftime('%d/%m/%Y'))  # Ngày chạy
        ws.cell(row=row, column=5, value=item['license_plate'])  # Biển số xe
        ws.cell(row=row, column=6, value=item['daily_salary'])  # Lương chuyến
    
    # Định dạng số cho cột lương
    for row in range(5, 5 + len(salary_data)):
        ws.cell(row=row, column=6).number_format = '#,##0'  # Định dạng số VNĐ
    
    # Dòng tổng cộng
    if salary_data:
        total_row = 5 + len(salary_data)
        total_salary = sum(item['daily_salary'] for item in salary_data)
        
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_salary).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0'
    
    # Điều chỉnh độ rộng cột
    column_widths = [8, 25, 15, 15, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    filename = f"BangTinhLuong_{month:02d}_{year}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/finance-report", response_class=HTMLResponse)
async def finance_report_page(
    request: Request, 
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None
):
    # Mặc định là tháng hiện tại nếu không có tham số
    if not month or not year:
        current_date = datetime.now()
        month = month or current_date.month
        year = year or current_date.year
    
    # Lấy dữ liệu tài chính từ bảng FinanceTransaction riêng biệt
    finance_data = db.query(FinanceTransaction).filter(
        and_(
            extract('month', FinanceTransaction.date) == month,
            extract('year', FinanceTransaction.date) == year
        )
    ).order_by(FinanceTransaction.date.desc()).all()
    
    # Tính tổng từ bảng mới
    total_income = sum(item.total for item in finance_data if item.transaction_type == "Thu")
    total_expense = sum(item.total for item in finance_data if item.transaction_type == "Chi")
    total_balance = total_income - total_expense
    
    return templates.TemplateResponse("finance_report.html", {
        "request": request,
        "finance_data": finance_data,
        "total_income": total_income,
        "total_expense": total_expense,
        "total_balance": total_balance,
        "selected_month": month,
        "selected_year": year
    })

@app.get("/finance-report/export")
async def export_finance_report_excel(
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None
):
    # Mặc định là tháng hiện tại nếu không có tham số
    if not month or not year:
        current_date = datetime.now()
        month = month or current_date.month
        year = year or current_date.year
    
    # Lấy dữ liệu tài chính từ bảng FinanceTransaction riêng biệt
    finance_data = db.query(FinanceTransaction).filter(
        and_(
            extract('month', FinanceTransaction.date) == month,
            extract('year', FinanceTransaction.date) == year
        )
    ).order_by(FinanceTransaction.date).all()
    
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"BaoCaoTaiChinh_{month:02d}_{year}"
    
    # Tiêu đề
    ws.cell(row=1, column=1, value=f"BÁO CÁO TÀI CHÍNH THÁNG {month}/{year}").font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Header bảng
    headers = ["Ngày", "Danh mục", "Diễn giải", "Chi", "Thu", "Thành tiền"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Dữ liệu
    for row, item in enumerate(finance_data, 4):
        ws.cell(row=row, column=1, value=item.date.strftime('%d/%m/%Y') if item.date else '')
        ws.cell(row=row, column=2, value=item.category or '')
        ws.cell(row=row, column=3, value=item.description or '')
        ws.cell(row=row, column=4, value=item.expense if item.expense > 0 else '')
        ws.cell(row=row, column=5, value=item.income if item.income > 0 else '')
        ws.cell(row=row, column=6, value=item.balance if item.balance else '')
        
        # Định dạng số cho các cột tiền
        for col in [4, 5, 6]:
            ws.cell(row=row, column=col).number_format = '#,##0'
    
    # Dòng tổng cộng
    if finance_data:
        total_row = 4 + len(finance_data)
        total_income = sum(item.income for item in finance_data)
        total_expense = sum(item.expense for item in finance_data)
        total_balance = total_income - total_expense
        
        ws.cell(row=total_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value="").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_expense).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_income).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_balance).font = Font(bold=True)
        
        # Định dạng số cho dòng tổng
        for col in [4, 5, 6]:
            ws.cell(row=total_row, column=col).number_format = '#,##0'
    
    # Điều chỉnh độ rộng cột
    column_widths = [12, 15, 30, 15, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Lưu vào memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Tạo tên file
    filename = f"BaoCaoTaiChinh_{month:02d}_{year}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/finance-report/create-sample-data")
async def create_sample_finance_data(db: Session = Depends(get_db)):
    """Tạo dữ liệu mẫu cho báo cáo tài chính"""
    current_date = datetime.now()
    
    # Dữ liệu mẫu cho tháng hiện tại
    sample_data = [
        {
            "date": current_date.replace(day=1),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 5000000,
            "expense": 0,
            "balance": 5000000
        },
        {
            "date": current_date.replace(day=2),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-12345",
            "income": 0,
            "expense": 2000000,
            "balance": -2000000
        },
        {
            "date": current_date.replace(day=3),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 4500000,
            "expense": 0,
            "balance": 4500000
        },
        {
            "date": current_date.replace(day=5),
            "category": "Chi",
            "description": "Chi phí sửa chữa xe 51A-67890",
            "income": 0,
            "expense": 1500000,
            "balance": -1500000
        },
        {
            "date": current_date.replace(day=10),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 4800000,
            "expense": 0,
            "balance": 4800000
        },
        {
            "date": current_date.replace(day=12),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-12345",
            "income": 0,
            "expense": 1800000,
            "balance": -1800000
        },
        {
            "date": current_date.replace(day=15),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 5200000,
            "expense": 0,
            "balance": 5200000
        },
        {
            "date": current_date.replace(day=18),
            "category": "Chi",
            "description": "Chi phí bảo hiểm xe",
            "income": 0,
            "expense": 3000000,
            "balance": -3000000
        },
        {
            "date": current_date.replace(day=20),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_002",
            "income": 4600000,
            "expense": 0,
            "balance": 4600000
        },
        {
            "date": current_date.replace(day=25),
            "category": "Chi",
            "description": "Chi phí đổ dầu xe 51A-67890",
            "income": 0,
            "expense": 2200000,
            "balance": -2200000
        },
        {
            "date": current_date.replace(day=28),
            "category": "Thu",
            "description": "Thu tiền vận chuyển tuyến NA_004",
            "income": 5100000,
            "expense": 0,
            "balance": 5100000
        },
        {
            "date": current_date.replace(day=30),
            "category": "Chi",
            "description": "Chi phí lương lái xe",
            "income": 0,
            "expense": 8000000,
            "balance": -8000000
        }
    ]
    
    # Xóa dữ liệu cũ nếu có
    db.query(FinanceTransaction).delete()
    
    # Thêm dữ liệu mẫu vào bảng mới
    for data in sample_data:
        # Chuyển đổi dữ liệu từ format cũ sang format mới
        transaction = FinanceTransaction(
            transaction_type=data["category"],
            category=data["category"],
            date=data["date"],
            description=data["description"],
            route_code=data.get("route_code", ""),
            amount=data["amount_before_vat"],
            vat=data["vat_rate"],
            discount1=data["discount1_rate"],
            discount2=data["discount2_rate"],
            total=data["final_amount"],
            note=data.get("notes", ""),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        db.add(transaction)
    
    db.commit()
    
    return JSONResponse({
        "message": f"Đã tạo {len(sample_data)} bản ghi tài chính mẫu cho tháng {current_date.month}/{current_date.year}",
        "count": len(sample_data)
    })

@app.post("/finance-report/add")
async def add_finance_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Thêm bản ghi tài chính mới"""
    try:
        form_data = await request.form()
        
        # Lấy dữ liệu từ form
        date_str = form_data.get("date")
        category = form_data.get("category")
        description = form_data.get("description")
        route_code = form_data.get("route_code", "")
        
        # Xử lý các trường số, đảm bảo không bị lỗi khi chuỗi rỗng
        amount_before_vat_str = form_data.get("amount_before_vat", "0")
        vat_rate_str = form_data.get("vat_rate", "0")
        discount1_rate_str = form_data.get("discount1_rate", "0")
        discount2_rate_str = form_data.get("discount2_rate", "0")
        
        # Convert sang float, xử lý trường hợp chuỗi rỗng
        amount_before_vat = float(amount_before_vat_str) if amount_before_vat_str else 0.0
        vat_rate = float(vat_rate_str) if vat_rate_str else 0.0
        discount1_rate = float(discount1_rate_str) if discount1_rate_str else 0.0
        discount2_rate = float(discount2_rate_str) if discount2_rate_str else 0.0
        
        notes = form_data.get("notes", "")
        
        # Parse ngày
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Tính thành tiền theo công thức
        # Thành tiền = Số tiền + (Số tiền * VAT/100) - (Số tiền * CK1/100) - (Số tiền * CK2/100)
        vat_amount = amount_before_vat * (vat_rate / 100)
        discount1_amount = amount_before_vat * (discount1_rate / 100)
        discount2_amount = amount_before_vat * (discount2_rate / 100)
        final_amount = amount_before_vat + vat_amount - discount1_amount - discount2_amount
        
        # Tạo bản ghi mới trong bảng FinanceTransaction riêng biệt
        finance_transaction = FinanceTransaction(
            transaction_type=category,  # Thu/Chi
            category=category,  # Danh mục
            date=date_obj,  # Ngày thu/chi
            description=description,  # Diễn giải
            route_code=route_code,  # Mã tuyến (nếu có)
            amount=amount_before_vat,  # Số tiền chưa VAT
            vat=vat_rate,  # VAT (%)
            discount1=discount1_rate,  # Chiết khấu 1 (%)
            discount2=discount2_rate,  # Chiết khấu 2 (%)
            total=final_amount,  # Thành tiền
            note=notes,  # Ghi chú
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        db.add(finance_transaction)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã thêm bản ghi tài chính thành công",
            "record_id": finance_transaction.id
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi thêm bản ghi: {str(e)}"
        }, status_code=400)

@app.get("/finance-report/get/{record_id}")
async def get_finance_record(record_id: int, db: Session = Depends(get_db)):
    """Lấy thông tin bản ghi tài chính theo ID"""
    try:
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        return JSONResponse({
            "success": True,
            "data": {
                "id": finance_record.id,
                "transaction_type": finance_record.transaction_type,
                "date": finance_record.date.strftime("%Y-%m-%d") if finance_record.date else None,
                "description": finance_record.description,
                "route_code": finance_record.route_code,
                "amount": finance_record.amount,
                "vat": finance_record.vat,
                "discount1": finance_record.discount1,
                "discount2": finance_record.discount2,
                "total": finance_record.total,
                "note": finance_record.note
            }
        })
        
    except Exception as e:
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi lấy thông tin bản ghi: {str(e)}"
        }, status_code=500)

@app.post("/finance-report/edit")
async def edit_finance_record(
    request: Request,
    db: Session = Depends(get_db)
):
    """Sửa bản ghi tài chính"""
    try:
        form_data = await request.form()
        
        # Lấy ID bản ghi cần sửa
        record_id = form_data.get("record_id")
        if not record_id:
            return JSONResponse({
                "success": False,
                "message": "Thiếu ID bản ghi"
            }, status_code=400)
        
        # Tìm bản ghi trong database
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        # Lấy dữ liệu từ form
        date_str = form_data.get("date")
        category = form_data.get("category")
        description = form_data.get("description")
        route_code = form_data.get("route_code", "")
        
        # Xử lý các trường số, đảm bảo không bị lỗi khi chuỗi rỗng
        amount_before_vat_str = form_data.get("amount_before_vat", "0")
        vat_rate_str = form_data.get("vat_rate", "0")
        discount1_rate_str = form_data.get("discount1_rate", "0")
        discount2_rate_str = form_data.get("discount2_rate", "0")
        
        # Convert sang float, xử lý trường hợp chuỗi rỗng
        amount_before_vat = float(amount_before_vat_str) if amount_before_vat_str else 0.0
        vat_rate = float(vat_rate_str) if vat_rate_str else 0.0
        discount1_rate = float(discount1_rate_str) if discount1_rate_str else 0.0
        discount2_rate = float(discount2_rate_str) if discount2_rate_str else 0.0
        
        notes = form_data.get("notes", "")
        
        # Parse ngày
        from datetime import datetime
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        # Tính thành tiền theo công thức
        vat_amount = amount_before_vat * (vat_rate / 100)
        discount1_amount = amount_before_vat * (discount1_rate / 100)
        discount2_amount = amount_before_vat * (discount2_rate / 100)
        final_amount = amount_before_vat + vat_amount - discount1_amount - discount2_amount
        
        # Cập nhật bản ghi
        finance_record.transaction_type = category
        finance_record.category = category
        finance_record.date = date_obj
        finance_record.description = description
        finance_record.route_code = route_code
        finance_record.amount = amount_before_vat
        finance_record.vat = vat_rate
        finance_record.discount1 = discount1_rate
        finance_record.discount2 = discount2_rate
        finance_record.total = final_amount
        finance_record.note = notes
        finance_record.updated_at = datetime.utcnow()
        
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã cập nhật bản ghi tài chính thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi cập nhật bản ghi: {str(e)}"
        }, status_code=400)

@app.delete("/finance-report/delete/{record_id}")
async def delete_finance_record(record_id: int, db: Session = Depends(get_db)):
    """Xóa bản ghi tài chính"""
    try:
        finance_record = db.query(FinanceTransaction).filter(FinanceTransaction.id == record_id).first()
        
        if not finance_record:
            return JSONResponse({
                "success": False,
                "message": "Không tìm thấy bản ghi tài chính"
            }, status_code=404)
        
        db.delete(finance_record)
        db.commit()
        
        return JSONResponse({
            "success": True,
            "message": "Đã xóa bản ghi tài chính thành công"
        })
        
    except Exception as e:
        db.rollback()
        return JSONResponse({
            "success": False,
            "message": f"Lỗi khi xóa bản ghi: {str(e)}"
        }, status_code=500)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
